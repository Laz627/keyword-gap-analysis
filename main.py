import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Color, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.formatting.rule import CellIsRule
from urllib.parse import urlparse

# Streamlit app configuration
st.set_page_config(page_title="Keyword Ranking Analysis", layout="wide")
st.title("Keyword Ranking Analysis")

# Sidebar inputs
st.sidebar.header("Target Domain")
target_domain = st.sidebar.text_input("Enter the target domain (e.g., 'pella.com')")

st.sidebar.header("Data Upload")
uploaded_files = st.sidebar.file_uploader("Upload CSV Files", type=["csv"], accept_multiple_files=True)

def extract_domain(url):
    """Extract domain name from URL"""
    try:
        if pd.isna(url) or not url:
            return None
        url = url.lower()
        # Add protocol if not present
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        parsed_url = urlparse(url)
        domain = parsed_url.netloc
        if domain.startswith('www.'):
            domain = domain[4:]
        return domain
    except:
        return url  # Return original if parsing fails

def process_competitor_data(data_frames, target_domain):
    """Process and combine competitor data with specific column ordering"""
    # Initialize empty DataFrame for final results
    all_keywords = set()
    search_volumes = {}
    domains = {}
    
    # First pass to collect all keywords and search volumes
    for df in data_frames:
        df.columns = [col.strip() for col in df.columns]
        all_keywords.update(df['Keyword'].unique())
        for keyword, volume in zip(df['Keyword'], df['Search Volume']):
            if keyword not in search_volumes or (volume > 0 and search_volumes[keyword] == 0):
                search_volumes[keyword] = volume
    
    final_df = pd.DataFrame(list(all_keywords), columns=['Keyword'])
    final_df['Search Volume'] = final_df['Keyword'].map(search_volumes)
    
    # Initialize target domain columns
    target_domain_name = target_domain
    final_df[f'{target_domain_name} Rank'] = 100  # Default value
    final_df[f'{target_domain_name} URL'] = None
    
    # Process each dataset
    competitor_count = 0
    competitor_domains = []
    
    for i, df in enumerate(data_frames):
        temp_df = df[['Keyword', 'URL', 'Rank']].copy()
        
        # Extract domains for all URLs in this dataset
        domains_in_df = temp_df['URL'].apply(lambda x: extract_domain(x) if pd.notnull(x) else None)
        unique_domains = domains_in_df.dropna().unique()
        
        # Check if this dataset contains the target domain
        contains_target = domains_in_df.apply(lambda x: target_domain.lower() in str(x).lower() if pd.notnull(x) else False)
        
        if contains_target.any():
            # This contains target domain data
            target_data = temp_df[contains_target].copy()
            
            # Use the actual domain name for the column
            target_domain_name = target_domain
            
            final_df = final_df.merge(
                target_data[['Keyword', 'Rank', 'URL']], 
                on='Keyword', 
                how='left',
                suffixes=('', '_target')
            )
            final_df[f'{target_domain_name} Rank'] = final_df['Rank'].fillna(100)
            final_df[f'{target_domain_name} URL'] = final_df['URL']
            final_df.drop(['Rank', 'URL'], axis=1, inplace=True)
        else:
            # This is competitor data
            # Find domains that aren't the target
            competitor_domains_in_df = [d for d in unique_domains if target_domain.lower() not in str(d).lower()]
            
            if competitor_domains_in_df:
                for comp_domain in competitor_domains_in_df:
                    competitor_count += 1
                    competitor_domains.append(comp_domain)
                    
                    # Filter for rows with this competitor domain
                    comp_rows = domains_in_df == comp_domain
                    comp_data = temp_df[comp_rows].copy()
                    
                    if not comp_data.empty:
                        final_df = final_df.merge(
                            comp_data[['Keyword', 'Rank', 'URL']], 
                            on='Keyword', 
                            how='left',
                            suffixes=('', f'_comp{competitor_count}')
                        )
                        final_df[f'{comp_domain} Rank'] = final_df['Rank'].fillna(100)
                        final_df[f'{comp_domain} URL'] = final_df['URL']
                        final_df.drop(['Rank', 'URL'], axis=1, inplace=True)
    
    # Generate recommendations based on the updated rank thresholds
    def generate_recommendations(row):
        target_rank = row[f'{target_domain_name} Rank']
        
        if target_rank <= 3:
            return "Defend"
        elif 4 <= target_rank <= 10:
            return "Overtake"
        elif 11 <= target_rank <= 20:
            return "Optimize"
        elif 21 <= target_rank <= 40:
            return "Larger Adjustments"
        else:  # target_rank > 40 or target_rank == 100
            return "Create New"
    
    final_df['Recommendation'] = final_df.apply(generate_recommendations, axis=1)
    
    # Reorder columns
    rank_cols = [f'{target_domain_name} Rank'] + [f'{domain} Rank' for domain in competitor_domains]
    url_cols = [f'{target_domain_name} URL'] + [f'{domain} URL' for domain in competitor_domains]
    
    final_df = final_df[['Keyword', 'Recommendation', 'Search Volume'] + rank_cols + url_cols]
    
    return final_df, target_domain_name, competitor_domains

def calculate_average_rank(series):
    """Calculate average rank excluding 100 values (which represent N/A)"""
    valid_ranks = series[series != 100]
    if len(valid_ranks) == 0:
        return "N/A"
    return int(round(valid_ranks.mean(), 0))  # Round to whole number

def get_top_keywords_by_category(df, domain, domains_dict):
    """Get top 50 keywords sorted by search volume and rank"""
    if domain == domains_dict['target']:
        # For target domain, sort by search volume within each recommendation category
        defend_kw = df[df['Recommendation'] == 'Defend'].sort_values(
            by='Search Volume', ascending=False).head(50)[['Keyword', 'Search Volume', f'{domain} Rank']]
        overtake_kw = df[df['Recommendation'] == 'Overtake'].sort_values(
            by='Search Volume', ascending=False).head(50)[['Keyword', 'Search Volume', f'{domain} Rank']]
        optimize_kw = df[df['Recommendation'] == 'Optimize'].sort_values(
            by='Search Volume', ascending=False).head(50)[['Keyword', 'Search Volume', f'{domain} Rank']]
        larger_adjust_kw = df[df['Recommendation'] == 'Larger Adjustments'].sort_values(
            by='Search Volume', ascending=False).head(50)[['Keyword', 'Search Volume', f'{domain} Rank']]
        create_kw = df[df['Recommendation'] == 'Create New'].sort_values(
            by='Search Volume', ascending=False).head(50)[['Keyword', 'Search Volume', f'{domain} Rank']]
        return {
            'Defend Keywords': defend_kw,
            'Overtake Keywords': overtake_kw,
            'Optimize Keywords': optimize_kw,
            'Larger Adjustments Keywords': larger_adjust_kw,
            'Create New Keywords': create_kw
        }
    else:
        # For competitors, first sort by rank, then by search volume
        comp_rank_col = f'{domain} Rank'
        return df[df[comp_rank_col] != 100].sort_values(
            by=[comp_rank_col, 'Search Volume'],  # Sort by rank first, then search volume
            ascending=[True, False]  # Ascending for rank (better ranks first), descending for search volume
        ).head(50)[['Keyword', 'Search Volume', comp_rank_col]]

if uploaded_files and target_domain:
    # Process the uploaded files
    data_frames = [pd.read_csv(file) for file in uploaded_files]
    
    # Combine competitor data side by side
    merged_df, target_domain_name, competitor_domains = process_competitor_data(data_frames, target_domain)
    
    # Create domains dictionary for reference
    domains_dict = {
        'target': target_domain_name,
        'competitors': competitor_domains
    }
    
    # Display filters
    st.subheader("Filters")
    col1, col2 = st.columns(2)
    with col1:
        keyword_filter = st.text_input("Filter by keyword:")
    with col2:
        recommendation_filter = st.selectbox(
            "Filter by recommendation:",
            ["All", "Defend", "Overtake", "Optimize", "Larger Adjustments", "Create New"]
        )
    
    filtered_df = merged_df.copy()
    if keyword_filter:
        filtered_df = filtered_df[filtered_df["Keyword"].str.contains(keyword_filter, case=False, na=False)]
    if recommendation_filter != "All":
        filtered_df = filtered_df[filtered_df["Recommendation"] == recommendation_filter]

    # Check if filtered DataFrame is empty
    if filtered_df.empty:
        st.warning("No data to display after applying filters.")
        st.stop()

    # Display the DataFrame with styling
    st.subheader("Keyword Rankings Table")

    try:
        display_df = filtered_df.copy()
        
        # Create styler object
        styler = display_df.style
        
        # Define custom CSS properties
        styler.set_properties(**{
            'background-color': '#f5f5f5',
            'color': '#333333',
            'border': '1px solid #e0e0e0',
            'padding': '8px',
            'text-align': 'left'
        })
        
        # Apply background gradient only to rank columns (excluding 100s)
        rank_columns = [col for col in display_df.columns if 'Rank' in col]
        for col in rank_columns:
            mask = display_df[col] != 100
            styler.background_gradient(
                cmap='RdYlGn_r',
                vmin=1,
                vmax=30,
                subset=pd.IndexSlice[mask, col],
                text_color_threshold=0.7
            )
        
        # Format columns
        format_dict = {
            'Search Volume': '{:,.0f}',  # Add thousands separator
            **{col: '{:.0f}' for col in rank_columns}  # Format rank columns as integers
        }
        styler.format(format_dict)
        
        # Add header styling
        styler.set_table_styles([
            {'selector': 'thead th', 
             'props': [('background-color', '#2c3e50'), 
                      ('color', 'white'),
                      ('font-weight', 'bold'),
                      ('padding', '12px')]},
            {'selector': 'tbody tr:nth-of-type(even)',
             'props': [('background-color', '#f8f9fa')]},
            {'selector': 'td', 
             'props': [('padding', '8px'),
                      ('border', '1px solid #dee2e6')]}
        ])
        
        # Display the styled DataFrame
        st.dataframe(styler, use_container_width=True, height=600)
    
    except Exception as e:
        st.error(f"Error applying styling: {str(e)}")
        st.dataframe(filtered_df, use_container_width=True)

    # Summary statistics
    st.subheader("Summary Statistics")
    
    # Create list of metrics
    metrics = [
        'Average Rank (excluding N/A)', 
        'Keywords in Top 10',
        'Keywords in Top 3',
        'Not Ranking (N/A)',
        'Defend Keywords',
        'Overtake Keywords',
        'Optimize Keywords',
        'Larger Adjustments Keywords',
        'Create New Keywords'
    ]
    
    # Create summary stats for target domain
    target_stats = [
        calculate_average_rank(filtered_df[f'{target_domain_name} Rank']),
        len(filtered_df[filtered_df[f'{target_domain_name} Rank'] <= 10]),
        len(filtered_df[filtered_df[f'{target_domain_name} Rank'] <= 3]),
        len(filtered_df[filtered_df[f'{target_domain_name} Rank'] == 100]),
        len(filtered_df[filtered_df['Recommendation'] == 'Defend']),
        len(filtered_df[filtered_df['Recommendation'] == 'Overtake']),
        len(filtered_df[filtered_df['Recommendation'] == 'Optimize']),
        len(filtered_df[filtered_df['Recommendation'] == 'Larger Adjustments']),
        len(filtered_df[filtered_df['Recommendation'] == 'Create New'])
    ]
    
    # Initialize competitor stats
    competitor_stats = {}
    for comp_domain in competitor_domains:
        comp_col = f'{comp_domain} Rank'
        competitor_stats[comp_domain] = [
            calculate_average_rank(filtered_df[comp_col]),
            len(filtered_df[filtered_df[comp_col] <= 10]),
            len(filtered_df[filtered_df[comp_col] <= 3]),
            len(filtered_df[filtered_df[comp_col] == 100]),
            "N/A",  # Use "N/A" instead of "-" to avoid type conversion issues
            "N/A", 
            "N/A", 
            "N/A", 
            "N/A"
        ]
    
    # Build summary DataFrame
    summary_data = {
        'Metric': metrics,
        target_domain_name: target_stats
    }
    for comp_domain, stats in competitor_stats.items():
        summary_data[comp_domain] = stats
    
    summary_stats = pd.DataFrame(summary_data)
    
    # Style the summary statistics
    st.dataframe(summary_stats, use_container_width=True)

    # Display top keywords
    st.subheader("Top Keywords by Search Volume")
    
    # Get top keywords for target domain
    target_top_kw = get_top_keywords_by_category(filtered_df, target_domain_name, domains_dict)
    
    # Display target domain top keywords by category
    for category, df in target_top_kw.items():
        st.write(f"\n{category}:")
        st.dataframe(df, use_container_width=True)
    
    # Get and display competitor top keywords
    for comp_domain in competitor_domains:
        if f'{comp_domain} Rank' in filtered_df.columns:
            st.write(f"\n{comp_domain} Top Keywords:")
            comp_top_kw = get_top_keywords_by_category(filtered_df, comp_domain, domains_dict)
            st.dataframe(comp_top_kw, use_container_width=True)

    # Export to Excel
    st.subheader("Export Data")
    if st.button("Download Excel"):
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # Write main rankings sheet
            filtered_df.to_excel(writer, index=False, sheet_name="Rankings")
            
            # Write summary statistics
            summary_stats.to_excel(writer, index=False, sheet_name="Summary")
            
            # Get top keywords
            target_top_kw = get_top_keywords_by_category(filtered_df, target_domain_name, domains_dict)
            competitor_top_kw = {}
            for comp_domain in competitor_domains:
                if f'{comp_domain} Rank' in filtered_df.columns:
                    competitor_top_kw[comp_domain] = get_top_keywords_by_category(filtered_df, comp_domain, domains_dict)
            
            # Prepare top keywords data for single write
            top_kw_rows = []
            
            # Add target domain keywords (sorted by search volume)
            for category, df in target_top_kw.items():
                # Add category header
                top_kw_rows.append(pd.DataFrame({'Category': [category]}))
                # Add data (already sorted by search volume)
                top_kw_rows.append(df)
                # Add blank row
                top_kw_rows.append(pd.DataFrame({'Category': ['']}))
            
            # Add competitor keywords (sorted by rank, then search volume)
            for comp_name, comp_df in competitor_top_kw.items():
                # Add competitor header
                top_kw_rows.append(pd.DataFrame({'Category': [f"{comp_name} Top Keywords"]}))
                # Add data (already sorted by rank and search volume)
                top_kw_rows.append(comp_df)
                # Add blank row
                top_kw_rows.append(pd.DataFrame({'Category': ['']}))
            
            # Combine all top keywords data
            top_kw_df = pd.concat(top_kw_rows, ignore_index=True)
            
            # Write top keywords
            top_kw_df.to_excel(writer, sheet_name="Top Keywords", index=False)
            
            # Get workbook and worksheets
            workbook = writer.book
            rankings_ws = workbook["Rankings"]
            summary_ws = workbook["Summary"]
            top_kw_ws = workbook["Top Keywords"]
            
            # Define styles
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin', color='DEE2E6'),
                right=Side(style='thin', color='DEE2E6'),
                top=Side(style='thin', color='DEE2E6'),
                bottom=Side(style='thin', color='DEE2E6')
            )
            
            # Function to apply basic styling to a worksheet
            def apply_basic_styling(ws):
                # Style headers
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left')
                
                # Style all cells
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left')
                
                # Freeze top row
                ws.freeze_panes = 'A2'
            
            # Apply basic styling to all sheets
            for ws in [rankings_ws, summary_ws, top_kw_ws]:
                apply_basic_styling(ws)
            
# Rankings sheet specific styling
            rank_columns = [col for col in filtered_df.columns if 'Rank' in col]
            for col in rank_columns:
                col_idx = filtered_df.columns.get_loc(col) + 1
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                
                # First add light gray for rank 100 (this should be applied first)
                rankings_ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{len(filtered_df) + 1}',
                    CellIsRule(
                        operator='equal',
                        formula=['100'],
                        stopIfTrue=True,
                        fill=PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    )
                )
                
                # Then add color scale for ranks 1-99 using fixed values
                rankings_ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{len(filtered_df) + 1}',
                    ColorScaleRule(
                        start_type='num',
                        start_value=1,
                        start_color='63BE7B',  # Green
                        mid_type='num',
                        mid_value=25,  # Midpoint at rank 25
                        mid_color='FFEB84',  # Yellow
                        end_type='num',
                        end_value=50,  # End at rank 50
                        end_color='F8696B'  # Red
                    )
                )

                # Add number formatting and ensure rank 100 is gray
                for cell in rankings_ws[col_letter][1:]:
                    cell.number_format = '0'
                    if cell.value == 100:
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Format numbers in all sheets
            for ws in [rankings_ws, summary_ws, top_kw_ws]:
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        column_header = ws.cell(1, cell.column).value
                        if 'Search Volume' in str(column_header):
                            cell.number_format = '#,##0'
                        elif 'Rank' in str(column_header):
                            cell.number_format = '0'
            
            # Auto-fit columns and wrap URLs
            for ws in [rankings_ws, summary_ws, top_kw_ws]:
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    is_url_column = 'URL' in str(column[0].value)
                    
                    if is_url_column:
                        # Set fixed width for URL columns and wrap text
                        ws.column_dimensions[column[0].column_letter].width = 50
                        for cell in column:
                            cell.alignment = Alignment(wrap_text=True)
                    else:
                        # Calculate max length for other columns
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 40)
            
            # Add alternating row colors
            for ws in [rankings_ws, summary_ws, top_kw_ws]:
                for row in range(2, ws.max_row + 1):
                    if row % 2 == 0:
                        for cell in ws[row]:
                            cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        buffer.seek(0)
        st.download_button(
            label="Download Excel File",
            data=buffer,
            file_name="keyword_rankings.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("Upload your data and specify the target domain to begin.")
